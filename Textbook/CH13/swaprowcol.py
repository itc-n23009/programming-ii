import openpyxl

def transpose_spreadsheet(input_file, output_file):
    wb = openpyxl.load_workbook(input_file)
    sheet = wb.active

    sheet_data = []
    for row in sheet.iter_rows(values_only=True):
        sheet_data.append(list(row))

    transposed_wb = openpyxl.Workbook()
    transposed_sheet = transposed_wb.active

    for i in range(len(sheet_data)):
        for j in range(len(sheet_data[i])):
            transposed_sheet.cell(row=j+1, column=i+1, value=sheet_data[i][j])

    transposed_wb.save(output_file)

input_file = 'input.xlsx'
output_file = 'output.xlsx'

transpose_spreadsheet(input_file, output_file)

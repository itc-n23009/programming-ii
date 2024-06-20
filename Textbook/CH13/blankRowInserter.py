import openpyxl, sys

def insert_blank_rows(filename, N, M):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active

    for row in ws.iter_rows(min_row=1, max_row=N, values_only=True):
        new_ws.append(row)
    
    for _ in range(M):
        new_ws.append([None] * ws.max_column)
    
    for row in ws.iter_rows(min_row=N+1, values_only=True):
        new_ws.append(row)
    
    new_filename = "modified" + filename
    new_wb.save(new_filename)
    print(f"Modified file has been saved as {new_filename}")

def main():
    if len(sys.argv) != 4:
        print("Usage: python3 blankRowInserter.py <N> <M> <filename>")
        sys.exit(1)

    try:
        N = int(sys.argv[1])
        M = int(sys.argv[2])
        filename = sys.argv[3]
    except ValueError:
        print("Please provide valid integers for N and M.")
        sys.exit(1)

    insert_blank_rows(filename, N, M)

if __name__ == "__main__":
    main()

